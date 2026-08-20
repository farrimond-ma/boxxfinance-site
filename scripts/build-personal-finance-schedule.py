#!/usr/bin/env python3
"""
Builds a 90-weekday personal-finance content schedule for the Boxx Finance
content engine — 8 new pillars (property/personal finance) designed to
drive Google Discover traffic and topical authority, additive to the
existing commercial-finance pillars.

Output: personal-finance-schedule-90day.xlsx (project root), 3 sheets:
  1. ContentEngine — one row per piece of content (2 blogs + 5 locations
     per weekday)
  2. Pillar Overview — one row per pillar, audience/keywords/intent/funnel
  3. Setup Notes — schedule totals, compliance notes, next steps

This is a PLANNING document for human review, not a live publish queue —
scripts/content-engine/sync-personal-finance-schedule.js reads this file
and pushes properly-shaped rows into the actual ContentEngine Google Sheet
once reviewed. See that script's header for the column-mapping rationale.

Run: python scripts/build-personal-finance-schedule.py [--start YYYY-MM-DD]
"""
import sys
import re
import datetime
import argparse

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NAVY = '010816'
GOLD = 'B8922A'

# ─── Pillars ──────────────────────────────────────────────────────────────
# Verbatim from the brief. Each pillar also carries the metadata needed for
# the Pillar Overview sheet and the funnel mapping used by the sync script —
# these personal-finance topics don't have a matching /funding-solutions/
# page of their own, so each pillar is mapped to whichever existing Boxx
# commercial service is the most credible next step for that reader.
PILLARS = {
    11: {
        'name': 'Buying a Home',
        'color': 'E3F2FD',
        'audience': 'First-time buyers and home movers researching the purchase process',
        'keywords': ['stamp duty calculator', 'conveyancing fees', 'house buying process UK', 'exchange of contracts', 'homebuyer survey', 'buying at auction'],
        'llm_questions': ['How much stamp duty will I pay?', 'What happens on completion day?', 'How long does conveyancing take?', 'How do I buy a house at auction?'],
        'search_intent': 'Informational, transactional-adjacent (early research stage)',
        'discover_potential': 'Very High',
        'funnel_service': 'Bridging Loans',
        'funnel_url': '/funding-solutions/bridging-loans',
        'funnel_reason': 'Chain breaks, auction purchases and completion-day timing pressure are exactly what bridging solves',
        'topics': [
            'stamp duty calculator UK', 'how much stamp duty will I pay', 'stamp duty for first time buyers',
            'stamp duty on second homes', 'how to avoid stamp duty legally', 'what are solicitor fees when buying a house',
            'conveyancing fees explained', 'how long does conveyancing take', 'what is exchange of contracts',
            'what happens on completion day', 'how to make an offer on a house', 'what is a homebuyer survey',
            'how to negotiate house price down', 'hidden costs of buying a home', 'how to buy a house at auction',
        ],
    },
    12: {
        'name': 'Remortgaging',
        'color': 'FDF3D7',
        'audience': 'Existing homeowners approaching the end of a fixed rate or wanting to release equity',
        'keywords': ['remortgage deals UK', 'remortgage to release equity', 'remortgage bad credit', 'product transfer vs remortgage', 'remortgage exit fees', 'self employed remortgage'],
        'llm_questions': ['When should I remortgage?', 'Can I remortgage with bad credit?', 'How long does remortgaging take?', 'Should I remortgage now or wait?'],
        'search_intent': 'Informational leaning transactional (active decision window)',
        'discover_potential': 'Very High',
        'funnel_service': 'Bridging Loans',
        'funnel_url': '/funding-solutions/bridging-loans',
        'funnel_reason': 'Readers stuck between a bad remortgage rate and a time-sensitive need are a natural bridging-to-remortgage audience',
        'topics': [
            'when should I remortgage', 'how to remortgage step by step', 'best remortgage deals UK',
            'remortgage to release equity', 'how much equity do I need to remortgage', 'can I remortgage with bad credit',
            'remortgage vs product transfer', 'how long does remortgaging take', 'remortgage early exit fees explained',
            'should I remortgage now or wait', 'remortgage to consolidate debt', 'remortgage for home improvements',
            'self employed remortgage UK', 'remortgage on interest only mortgage', 'what happens if I cant remortgage',
        ],
    },
    13: {
        'name': 'Bad Credit and Debt',
        'color': 'FCE4EC',
        'audience': 'Borrowers with credit problems trying to understand their options',
        'keywords': ['mortgage with bad credit', 'bad credit loans UK', 'improve credit score', 'CCJ mortgage', 'IVA explained', 'debt consolidation loans'],
        'llm_questions': ['Can I get a mortgage with bad credit?', 'What is a CCJ and how does it affect borrowing?', 'Can you remortgage with a CCJ?', 'How long does bad credit stay on your file?'],
        'search_intent': 'Informational, high anxiety/urgency',
        'discover_potential': 'High',
        'funnel_service': 'Bridging Loans',
        'funnel_url': '/funding-solutions/bridging-loans',
        'funnel_reason': 'Adverse-credit bridging is a genuine Boxx product for readers mainstream lenders have already declined',
        'topics': [
            'how to get a mortgage with bad credit', 'bad credit loans UK', 'how to improve your credit score fast',
            'what is a CCJ and how does it affect borrowing', 'IVA explained', 'debt consolidation loans UK',
            'how to get out of debt fast', 'debt management plan explained', 'what happens if you cant pay your mortgage',
            'can you remortgage with a CCJ', 'secured loans for bad credit', 'guarantor loans explained',
            'how to check your credit score for free', 'what is a default on your credit file', 'how long does bad credit stay on your file',
        ],
    },
    14: {
        'name': 'Property Investing for Beginners',
        'color': 'E8F5E9',
        'audience': 'Aspiring and early-stage landlords/investors researching how to start',
        'keywords': ['buy to let for beginners', 'rental yield calculator', 'best areas to invest UK', 'HMO investing', 'holiday let investing', 'property portfolio'],
        'llm_questions': ['How much deposit do I need for a buy to let?', 'How do I calculate rental yield?', 'Is buy to let still worth it in 2025?', 'What are common property investment mistakes?'],
        'search_intent': 'Informational, top-of-funnel',
        'discover_potential': 'Exceptional',
        'funnel_service': 'Bridging Loans',
        'funnel_url': '/funding-solutions/bridging-loans',
        'funnel_reason': 'Direct overlap — investors researching how to start are exactly Boxx\'s core bridging/refurbishment audience',
        'topics': [
            'how to start investing in property UK', 'buy to let for beginners', 'how much deposit do I need for a buy to let',
            'is buy to let still worth it in 2025', 'how to find the right buy to let property', 'best areas to invest in property UK',
            'how to calculate rental yield', 'gross yield vs net yield explained', 'what is capital growth in property',
            'how to build a property portfolio from scratch', 'HMO investing for beginners', 'holiday let investing UK',
            'flipping houses UK is it worth it', 'property investment vs stocks and shares', 'common property investment mistakes',
        ],
    },
    15: {
        'name': 'Saving and Building Wealth',
        'color': 'FFF8E1',
        'audience': 'General audience building financial literacy and long-term savings habits',
        'keywords': ['best savings accounts UK', 'cash ISA vs stocks and shares ISA', 'emergency fund', 'budgeting UK', 'lifetime ISA', 'retire early UK'],
        'llm_questions': ['How much should I have in savings?', 'What is the 50/30/20 budget rule?', 'Pension vs property investment — which is better?', 'How can I retire early in the UK?'],
        'search_intent': 'Informational, broad top-of-funnel',
        'discover_potential': 'High',
        'funnel_service': 'Funding Solutions (hub)',
        'funnel_url': '/funding-solutions',
        'funnel_reason': 'Weakest direct product overlap of the 8 pillars — funnels to the general hub rather than a specific product; valuable mainly for topical authority and Discover reach',
        'topics': [
            'how to save money fast UK', 'best savings accounts UK', 'cash ISA vs stocks and shares ISA',
            'how much should I have in savings', 'how to build an emergency fund', 'how to save for a house deposit',
            'how to save money on your mortgage', 'pension vs property investment', 'how to make your money work harder',
            'compound interest explained simply', 'how to budget effectively UK', '50 30 20 budget rule explained',
            'how to stop living paycheck to paycheck', 'lifetime ISA explained', 'how to retire early UK',
        ],
    },
    16: {
        'name': 'Legal and Conveyancing',
        'color': 'F3E5F5',
        'audience': 'Buyers/sellers mid-transaction with a specific legal question',
        'keywords': ['conveyancing solicitor', 'conveyancing delays', 'leasehold vs freehold', 'title deed', 'buyer pulled out', 'property through limited company'],
        'llm_questions': ['What can delay conveyancing?', 'What is a flying freehold?', 'What happens if a buyer pulls out?', 'Can I buy a property through a limited company?'],
        'search_intent': 'Informational, transaction-stage urgency',
        'discover_potential': 'High',
        'funnel_service': 'Bridging Loans',
        'funnel_url': '/funding-solutions/bridging-loans',
        'funnel_reason': 'Conveyancing delays and buyer-pulled-out scenarios are the exact trigger events Boxx\'s bridging content already targets',
        'topics': [
            'what is conveyancing', 'how much do solicitors charge for conveyancing', 'how to choose a conveyancing solicitor',
            'what does a conveyancer do', 'how long does conveyancing take UK', 'what can delay conveyancing',
            'what are searches in conveyancing', 'what is a title deed', 'leasehold vs freehold explained',
            'what is a flying freehold', 'what happens if a buyer pulls out', 'how to transfer property ownership',
            'what is probate property', 'buying a property through a limited company legally', 'legal fees for remortgaging explained',
        ],
    },
    17: {
        'name': 'Property News and Market Updates',
        'color': 'E0F2F1',
        'audience': 'News-driven readers reacting to a specific announcement or data release',
        'keywords': ['Bank of England rate decision', 'house prices UK', 'rental market crisis', 'landlord regulation 2025', 'housing shortage UK', 'mortgage rates forecast'],
        'llm_questions': ['What does the latest Bank of England decision mean for the property market?', 'Is now a good time to buy property in the UK?', 'What are the property market hotspots?', 'How is inflation affecting property prices?'],
        'search_intent': 'News/reactive — highest Discover potential of any pillar, but time-sensitive',
        'discover_potential': 'Exceptional',
        'funnel_service': 'Funding Solutions (hub)',
        'funnel_url': '/funding-solutions',
        'funnel_reason': 'Reactive news content — funnel varies by story, default to the general hub unless a specific article clearly maps to one product',
        'topics': [
            'Bank of England rate decision property market impact', 'house prices rising what it means for buyers',
            'house prices falling what it means for investors', 'rental market crisis explained', 'landlord regulation changes 2025',
            'best areas for property investment UK 2025', 'UK housing shortage explained', 'first time buyer schemes 2025 explained',
            'property market forecast 2025', 'what is happening to mortgage rates', 'government housing policy explained',
            'new build vs old build property investment', 'property market regional hotspots UK', 'impact of inflation on property prices',
            'is now a good time to buy property UK',
        ],
    },
    18: {
        'name': 'Taxes and Property',
        'color': 'EFEBE9',
        'audience': 'Landlords and property owners with a specific tax question',
        'keywords': ['capital gains tax property', 'Section 24 landlord tax', 'inheritance tax property', 'limited company property tax', 'rental income tax', 'stamp duty second homes'],
        'llm_questions': ['How do I reduce capital gains tax on property?', 'What is Section 24 and how does it affect landlords?', 'What expenses can landlords claim?', 'How do I minimise inheritance tax on a property portfolio?'],
        'search_intent': 'Informational, higher-value audience (existing property owners)',
        'discover_potential': 'High',
        'funnel_service': 'Commercial Mortgages',
        'funnel_url': '/funding-solutions/commercial-mortgages',
        'funnel_reason': 'Landlords restructuring for tax efficiency (e.g. moving to a limited company) commonly need to refinance at the same time',
        'topics': [
            'stamp duty explained in full', 'stamp duty on second homes and buy to let', 'how to reduce stamp duty legally',
            'capital gains tax on property UK', 'how to reduce capital gains tax on property', 'inheritance tax and property UK',
            'Section 24 tax explained for landlords', 'limited company tax advantages for property investors', 'what expenses can landlords claim',
            'income tax on rental income UK', 'how to declare rental income to HMRC', 'tax on holiday let income',
            'what is annual tax on enveloped dwellings', 'tax implications of selling a buy to let', 'how to minimise inheritance tax on a property portfolio',
        ],
    },
}

# ─── UK cities only ─────────────────────────────────────────────────────────
# Deliberately UK-only: stamp duty, conveyancing, ISAs, CCJs, Section 24 and
# remortgaging are UK-specific concepts, unlike the international-audience
# reasoning that applies to Boxx's commercial bridging/development pillars.
CITIES = [
    'London', 'Manchester', 'Birmingham', 'Leeds', 'Glasgow', 'Bristol', 'Edinburgh', 'Liverpool', 'Sheffield', 'Cardiff',
    'Leicester', 'Nottingham', 'Southampton', 'Newcastle', 'Brighton', 'Coventry', 'Reading', 'Derby', 'Stoke', 'Wolverhampton',
    'Oxford', 'Cambridge', 'Norwich', 'Plymouth', 'Exeter', 'York', 'Chester', 'Bath', 'Bournemouth', 'Portsmouth',
    'Milton Keynes', 'Luton', 'Swindon', 'Gloucester', 'Cheltenham', 'Worcester', 'Hereford', 'Lincoln', 'Peterborough', 'Ipswich',
    'Colchester', 'Southend', 'Basildon', 'Guildford', 'Maidstone', 'Canterbury', 'Crawley', 'Basingstoke', 'Salisbury', 'Eastbourne',
]

AUTHORS = ['Mark Higgins', 'Tara Jameson']
AUTHOR_EMAILS = {'Mark Higgins': 'mark@boxxfinance.co.uk', 'Tara Jameson': 'tara@boxxfinance.co.uk'}


def slugify(text):
    s = text.lower().strip()
    s = re.sub(r"[^a-z0-9\s-]", '', s)
    s = re.sub(r"\s+", '-', s)
    s = re.sub(r"-+", '-', s)
    return s.strip('-')


def titleize(topic):
    """Turns a lowercase search-phrase topic into a readable article title."""
    t = topic.strip()
    t = t[0].upper() + t[1:]
    for acronym in ['uk', 'iva', 'ccj', 'isa', 'hmo', 'hmrc']:
        t = re.sub(rf'\b{acronym}\b', acronym.upper(), t, flags=re.IGNORECASE)
    if not t.rstrip().endswith(('?', '.', '!')):
        if t.lower().startswith(('how ', 'what ', 'when ', 'should ', 'can ', 'is ', 'why ')):
            t += '?'
    return t


def meta_description(topic, pillar_name):
    base = f"{titleize(topic)} — a clear, practical guide from Boxx Finance."
    if len(base) > 155:
        base = base[:152].rsplit(' ', 1)[0] + '...'
    return base


def location_title(pillar_name, city):
    return f"{pillar_name} in {city} | Boxx Finance"


def location_meta(pillar_name, city):
    return f"Expert {pillar_name.lower()} advice in {city}. Speak to the Boxx Finance team today about your requirements."


def add_weekdays(start_date, n):
    """Yield n weekday dates starting from start_date (inclusive if it's a weekday)."""
    d = start_date
    count = 0
    while count < n:
        if d.weekday() < 5:  # Mon-Fri
            yield d
            count += 1
        d += datetime.timedelta(days=1)


def next_weekday_on_or_after(d):
    while d.weekday() >= 5:
        d += datetime.timedelta(days=1)
    return d


def build():
    parser = argparse.ArgumentParser()
    parser.add_argument('--start', default=None, help='Start date YYYY-MM-DD (default: next weekday)')
    args = parser.parse_args()

    if args.start:
        start_date = datetime.datetime.strptime(args.start, '%Y-%m-%d').date()
    else:
        # The original brief specified 2025-09-01, which is now in the past —
        # defaulting to the next weekday from today instead of a stale date.
        start_date = next_weekday_on_or_after(datetime.date.today() + datetime.timedelta(days=1))

    weekdays = list(add_weekdays(start_date, 90))

    pillar_numbers = list(PILLARS.keys())
    # Flatten all (pillar, topic) pairs per pillar, in order, for blog cycling
    blog_queue = []
    for pnum in pillar_numbers:
        for topic in PILLARS[pnum]['topics']:
            blog_queue.append((pnum, topic))
    # 90 weekdays * 2 blogs/day = 180 slots; 8 pillars * 15 topics = 120 topics.
    # Cycle the queue a second time (with a varied title prefix) once exhausted
    # rather than running out — flagged in Setup Notes as something to review
    # once real performance data exists.
    while len(blog_queue) < len(weekdays) * 2:
        blog_queue += blog_queue[:len(weekdays) * 2 - len(blog_queue)]

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1 — ContentEngine
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'ContentEngine'

    headers = ['ID', 'Type', 'Status', 'Publish Date', 'Slot', 'Pillar', 'City',
               'Primary Keyword', 'Topic / Angle', 'Title', 'Slug', 'URL',
               'Meta Title', 'Meta Description', 'Author']

    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = 'Boxx Finance — Personal Finance Content Schedule (90-Day, 8 Pillars, Google Discover Focus)'
    title_cell.font = Font(bold=True, color='FFFFFF', size=13)
    title_cell.fill = PatternFill('solid', fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = Font(bold=True, color=GOLD)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(horizontal='left', vertical='center')

    ws.freeze_panes = 'A3'

    row_num = 3
    next_id = 2000
    blog_idx = 0
    city_pillar_idx = 0  # cycles through (city, pillar) combos for locations
    loc_row_toggle = 0

    stats = {'blogs': 0, 'locations': 0, 'weekdays': len(weekdays)}

    # Build the (city, pillar) cycling order once: cities cycle fastest so
    # each city appears across many different pillars over the 90 days.
    city_pillar_pairs = []
    for i in range(len(weekdays) * 5):
        city = CITIES[i % len(CITIES)]
        pnum = pillar_numbers[(i // len(CITIES)) % len(pillar_numbers)]
        city_pillar_pairs.append((city, pnum))

    for day_idx, date in enumerate(weekdays):
        date_str = date.isoformat()

        # ── 2 blog posts/day ──────────────────────────────────────────────
        for slot_idx in range(2):
            pnum, topic = blog_queue[blog_idx]
            blog_idx += 1
            pillar = PILLARS[pnum]
            author = AUTHORS[slot_idx % 2]  # AM = Mark, PM = Tara
            slot_label = f'Blog {slot_idx + 1}'
            title = titleize(topic)
            slug = slugify(topic)
            url = f'/insights/{slug}'
            meta_title = f'{title} | Boxx Finance'
            meta_desc = meta_description(topic, pillar['name'])

            values = [next_id, 'Blog', 'scheduled', date_str, slot_label, pillar['name'], '',
                      topic, topic, title, slug, url, meta_title, meta_desc, author]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=v)
                cell.fill = PatternFill('solid', fgColor=pillar['color'])
            next_id += 1
            row_num += 1
            stats['blogs'] += 1

        # ── 5 location pages/day ──────────────────────────────────────────
        for loc_slot in range(5):
            city, pnum = city_pillar_pairs[day_idx * 5 + loc_slot]
            pillar = PILLARS[pnum]
            slot_label = f'Loc {loc_slot + 1}'
            keyword = f"{pillar['name'].lower()} {city.lower()}"
            slug = slugify(f"{pillar['name']}-{city}")
            url = f'/locations/{slug}'
            title = location_title(pillar['name'], city)
            meta_title = f'{pillar["name"]} in {city} | Boxx Finance'
            meta_desc = location_meta(pillar['name'], city)
            author = AUTHORS[loc_row_toggle % 2]
            loc_row_toggle += 1

            values = [next_id, 'Location', 'scheduled', date_str, slot_label, pillar['name'], f'{city}, UK',
                      keyword, f"{pillar['name']} — {city}", title, slug, url, meta_title, meta_desc, author]
            fill_color = 'F5F5F5' if loc_slot % 2 == 0 else 'FFFFFF'
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=v)
                cell.fill = PatternFill('solid', fgColor=fill_color)
            next_id += 1
            row_num += 1
            stats['locations'] += 1

    widths = [8, 10, 11, 13, 9, 28, 16, 32, 34, 42, 40, 34, 46, 50, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 2 — Pillar Overview
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Pillar Overview')
    ws2.merge_cells('A1:H1')
    note_cell = ws2['A1']
    note_cell.value = ("These 8 pillars are additive to the existing Boxx commercial finance pillars. "
                        "They target UK audiences only, as all topics (stamp duty, conveyancing, ISAs, CCJs, "
                        "Section 24) are UK-specific. Designed to drive Google Discover traffic and funnel a "
                        "wider personal finance audience into Boxx's commercial services.")
    note_cell.font = Font(italic=True, color=NAVY)
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws2.row_dimensions[1].height = 45

    headers2 = ['Pillar', 'Target Audience', 'Top 6 Google Keywords', 'Top 4 LLM Questions',
                'Search Intent', 'Discover Potential', 'Funnels To (Boxx Service)', 'Why This Funnel']
    for col_idx, h in enumerate(headers2, start=1):
        c = ws2.cell(row=2, column=col_idx, value=h)
        c.font = Font(bold=True, color=GOLD)
        c.fill = PatternFill('solid', fgColor=NAVY)
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws2.freeze_panes = 'A3'

    for i, pnum in enumerate(pillar_numbers):
        pillar = PILLARS[pnum]
        r = 3 + i
        values = [
            f"Pillar {pnum} — {pillar['name']}",
            pillar['audience'],
            ', '.join(pillar['keywords']),
            ' / '.join(pillar['llm_questions']),
            pillar['search_intent'],
            pillar['discover_potential'],
            pillar['funnel_service'],
            pillar['funnel_reason'],
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=col_idx, value=v)
            cell.fill = PatternFill('solid', fgColor=pillar['color'])
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    widths2 = [30, 40, 40, 55, 32, 16, 26, 45]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 3 — Setup Notes
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Setup Notes')
    ws3.column_dimensions['A'].width = 34
    ws3.column_dimensions['B'].width = 90

    notes = [
        ('Summary', '8 new personal finance pillars targeting UK audiences only, via Google Discover.'),
        ('Why UK only', 'Stamp duty, conveyancing, ISAs, CCJs, Section 24 and remortgaging are UK-specific '
                         'concepts with no relevance to international audiences.'),
        ('Publishing schedule', '2 blogs + 5 location pages per weekday.'),
        ('90-day totals', f"{stats['blogs']} blog posts, {stats['locations']} location pages, "
                           f"{stats['blogs'] + stats['locations']} total pieces across {stats['weekdays']} weekdays."),
        ('Annualised projection (weekdays only, ~252/yr)',
         f"~{round(stats['blogs'] / stats['weekdays'] * 252)} blog posts, "
         f"~{round(stats['locations'] / stats['weekdays'] * 252)} location pages per year at this cadence."),
        ('Schedule start date', f"{weekdays[0].isoformat()} — the original brief said 2025-09-01, which had "
                                 "already passed by the time this was generated, so it defaulted to the next "
                                 "weekday instead. Pass --start YYYY-MM-DD to override."),
        ('Topic/slug reuse note', '120 unique blog topics (8 pillars x 15) against 180 blog slots (90 weekdays x 2), '
                              'and 400 unique pillar x city combinations (8 x 50) against 450 location slots '
                              '(90 x 5), both mean this sheet has some repeated slugs by design — about 520 of '
                              'the 630 rows are actually unique. sync-personal-finance-schedule.js dedupes by '
                              'slug before pushing to the live sheet, so only the first occurrence of each '
                              'repeated topic/location actually gets scheduled — the rest are silently skipped, '
                              'meaning some weekday slots in this 90-day plan won\'t produce a genuinely new page. '
                              'Fine for a first pass; worth writing a second batch of real topics (or more '
                              'cities) before relying on the schedule filling every single slot.'),
        ('Pillar 17 — Property News', 'This pillar is reactive by nature. Recommend triggering it manually '
                                       '(via a dedicated script, not the scheduled queue) within hours of a real '
                                       'news event, rather than publishing pre-written "news" on a fixed schedule '
                                       'months in advance, which reads as stale by the time it goes live. Set up '
                                       'Google Alerts for: Bank of England interest rate, UK house prices, stamp '
                                       'duty changes, landlord regulation, rental market UK, budget property.'),
        ('FCA compliance', 'All content must be informational only — no specific product or lender '
                            'recommendations without FCA authorisation. Every article should include a clear '
                            'disclaimer and end with a recommendation to speak to a qualified adviser. This is '
                            'especially important for Pillars 13 (Bad Credit and Debt) and 18 (Taxes and Property), '
                            'which stray closest to regulated advice territory.'),
        ('Fabricated-content guardrail', 'Boxx\'s existing bridging content had real problems with invented '
                                          'statistics, fabricated "in our experience" claims, and jurisdiction-blind '
                                          'tax language (fixed 2026-08). Apply the same discipline here from day '
                                          'one: no invented tax rates/thresholds without a real source, no '
                                          '"our clients tell us..." claims that aren\'t real, and property-tax '
                                          'content (Pillar 18 especially) must stay UK-nation-aware — SDLT '
                                          '(England/NI), LBTT (Scotland) and LTT (Wales) are different taxes with '
                                          'different rates, not one UK-wide "stamp duty".'),
        ('Launch cadence', 'Built at the full spec\'d volume (7 pieces/weekday) as requested. Worth noting: the '
                            'existing bridging vertical was deliberately cut from 14/week to 5/week this same week '
                            'specifically because high-volume generation was producing near-duplicate, thin, '
                            'fabricated-sounding content. Recommend watching the first 2-3 weeks of personal-finance '
                            'output closely before assuming this cadence is sustainable long-term.'),
        ('Referral opportunity', 'Consider partnering with an FCA-authorised residential mortgage broker for lead '
                                  'referrals on the regulated-advice-adjacent topics. Referral fees typically '
                                  '£200-£800 per completed case. Boxx generates the lead; the partner handles '
                                  'regulated advice.'),
        ('Authors', 'Mark Higgins (mark@boxxfinance.co.uk), Tara Jameson (tara@boxxfinance.co.uk). The original '
                     'brief also listed Andrew Farrimond — deliberately dropped here, since he was identified '
                     'and removed as a fabricated author sitewide earlier this same week.'),
        ('Domain', 'boxxfinance.co.uk'),
        ('Getting this live', 'This spreadsheet is a planning document, not a live publish queue. Once reviewed, '
                               'run scripts/content-engine/sync-personal-finance-schedule.js to push these rows '
                               'into the real ContentEngine Google Sheet — and note that the current blog-'
                               'publishing workflow only processes rows where service = "Bridging Finance", so it '
                               'won\'t pick these up until that\'s addressed too (see the sync script\'s own notes).'),
    ]

    r = 1
    for label, value in notes:
        ws3.cell(row=r, column=1, value=label).font = Font(bold=True, color=NAVY)
        cell = ws3.cell(row=r, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[r].height = max(15, 15 * (len(value) // 90 + 1))
        r += 1

    out_path = 'personal-finance-schedule-90day.xlsx'
    wb.save(out_path)

    print(f"Total rows generated: {stats['blogs'] + stats['locations']}")
    print(f"Blog posts count: {stats['blogs']}")
    print(f"Location pages count: {stats['locations']}")
    print(f"Weekdays covered: {stats['weekdays']} ({weekdays[0].isoformat()} to {weekdays[-1].isoformat()})")
    print(f"Pillars covered: {len(PILLARS)}")
    print(f"File path: {out_path}")


if __name__ == '__main__':
    build()
